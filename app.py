#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""赞美诗资料智能整理中心 · 本地服务（零依赖，离线可用）"""
from __future__ import annotations

import io
import json
import mimetypes
import os
import re
import secrets
import shutil
import sys
import threading
import time
import traceback
import urllib.parse
import uuid
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
UPLOADS = DATA / "uploads"
EXPORTS = DATA / "exports"
STATIC = ROOT / "static"
SAMPLES = ROOT / "samples"
DB_PATH = DATA / "db.json"

for d in (DATA, UPLOADS, EXPORTS, SAMPLES):
    d.mkdir(parents=True, exist_ok=True)

CATEGORIES = None
_cat_path = DATA / "categories.json"
if _cat_path.exists():
    try:
        CATEGORIES = json.loads(_cat_path.read_text("utf-8"))
    except Exception:
        CATEGORIES = None

ADMIN = None
_admin_path = DATA / "admin.json"
if _admin_path.exists():
    try:
        ADMIN = json.loads(_admin_path.read_text("utf-8"))
    except Exception:
        ADMIN = None
_ADMIN_TOKENS = {}  # token -> 过期时间戳


def _admin_enabled():
    return bool(ADMIN and ADMIN.get("hash"))


def _admin_token():
    import hashlib
    return hashlib.sha256(("hymn-admin-token:" + (ADMIN.get("hash") or "")).encode("utf-8")).hexdigest()[:32]


def _check_password(pwd):
    if not _admin_enabled():
        return True
    import hashlib
    return hashlib.sha256(((ADMIN.get("salt") or "") + pwd).encode("utf-8")).hexdigest() == ADMIN.get("hash")


def _check_auth(headers, qs=None):
    if not _admin_enabled():
        return True
    token = headers.get("X-Admin-Token") or ""
    if not token and qs:
        token = (qs.get("token") or [""])[0]
    if not token:
        # cookie
        ck = headers.get("Cookie") or ""
        m = re.search(r"hymn_admin=([^;]+)", ck)
        token = m.group(1) if m else ""
    if token == _admin_token():
        return True
    return token in _ADMIN_TOKENS and _ADMIN_TOKENS[token] > time.time()


def _is_public(path, method):
    if path.startswith("/static/") or path.startswith("/files/samples/"):
        return True
    if path in ("/", "/index.html", "/mobile", "/m", "/mobile.html", "/qr", "/phone",
                "/api/bootstrap", "/api/songs/template", "/api/auth/check"):
        return True
    if method == "POST" and path in ("/api/ocr", "/api/import", "/api/login"):
        return True
    if method == "POST" and path == "/api/songs":
        return True  # 手机端保存
    return False

PORT = int(os.environ.get("PORT", "8787"))
HOST = os.environ.get("HOST", "127.0.0.1")
sys.path.insert(0, str(ROOT))

import engine
import exporters
import parsers

STATUS_LABELS = exporters.STATUS_LABELS
STATUSES = list(STATUS_LABELS.keys())


def now_str():
    return time.strftime("%Y-%m-%dT%H:%M:%S")


def get_lan_ips():
    """获取本机局域网 IPv4 地址（UDP 连接法，不发送数据）。"""
    import socket
    ips = []
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        if ip and not ip.startswith("127."):
            ips.append(ip)
    except Exception:
        pass
    if not ips:
        try:
            import subprocess
            for iface in ("en0", "en1"):
                out = subprocess.run(["ipconfig", "getifaddr", iface], capture_output=True, text=True, timeout=3)
                ip = out.stdout.strip()
                if ip and not ip.startswith("127."):
                    ips.append(ip)
        except Exception:
            pass
    return ips


def mask_key(k):
    if not k:
        return ""
    return k[:6] + "…" + k[-4:] if len(k) > 12 else "****"


# ---------------------------------------------------------------- 存储
class Store:
    def __init__(self, path):
        self.path = Path(path)
        self.lock = threading.RLock()
        self.data = self._load()

    def _load(self):
        if self.path.exists():
            try:
                return json.loads(self.path.read_text("utf-8"))
            except Exception:
                pass
        return {"songs": [], "seq": 0, "imports": [], "exports": [], "settings": {}}

    def save(self):
        with self.lock:
            tmp = self.path.with_suffix(".tmp")
            tmp.write_text(json.dumps(self.data, ensure_ascii=False, indent=1), "utf-8")
            tmp.replace(self.path)

    def songs(self):
        return self.data["songs"]

    def next_seq(self):
        with self.lock:
            self.data["seq"] += 1
            return self.data["seq"]

    def find(self, sid):
        for s in self.songs():
            if s["id"] == sid:
                return s
        return None


def default_song(**kw):
    s = {
        "id": "", "number": "", "title": "", "firstLine": "", "lyricist": "", "composer": "",
        "translator": "", "tune": "", "key": "", "meter": "", "source": "", "comment": "",
        "lyrics": "", "themes": [], "scenarios": [], "musicTypes": [],
        "category": "", "subcategory": "", "uploader": "",
        "difficulty": 3, "difficultyStars": "★★★☆☆", "singability": 3, "singabilityStars": "★★★☆☆",
        "rating": 0, "status": "pending", "needsReview": False, "aiConfidence": 0,
        "flags": [], "dupGroup": None, "dupMatches": [], "dupResolved": False,
        "importBatch": "", "importSource": "手工添加", "createdAt": now_str(), "updatedAt": now_str(),
        "attachments": [], "ocrNote": "",
    }
    s.update({k: v for k, v in kw.items() if v is not None})
    return s


def classify_song(song):
    res = engine.classify(song, CATEGORIES)
    song.update({k: v for k, v in res.items() if k != "needsReview"})
    song["needsReview"] = bool(res["needsReview"] or song.get("needsReview"))
    return res


def song_flags(song):
    flags = []
    if song.get("needsReview"):
        flags.append("需人工确认")
    if song.get("dupGroup"):
        flags.append("疑似重复")
    if not (song.get("title") or "").strip() or not (song.get("lyrics") or "").strip():
        flags.append("资料不完整")
    if song.get("ocrNote"):
        flags.append(song["ocrNote"])
    song["flags"] = flags


# ---------------------------------------------------------------- 去重
def apply_dedup(store):
    songs = store.songs()
    groups = engine.find_duplicates(songs)
    for s in songs:
        s["dupGroup"] = None
        s["dupMatches"] = []
    for g in groups:
        for m in g["members"]:
            s = store.find(m["id"])
            if s:
                s["dupGroup"] = g["group"]
                s["dupMatches"] = [x for x in g["members"] if x["id"] != s["id"]]
    return groups


def resolve_duplicate(store, group, action, keep_id=None):
    songs = store.songs()
    members = [s for s in songs if s.get("dupGroup") == group and s.get("status") != "merged"]
    if action == "merge":
        keep = store.find(keep_id) if keep_id else (max(members, key=lambda s: len(s.get("lyrics") or "")) if members else None)
        if not keep:
            return {"ok": False, "msg": "未找到保留曲目"}
        for s in members:
            if s["id"] == keep["id"]:
                continue
            # 合并：填充空缺字段 + 拼接歌词 + 迁移附件
            for f in ("lyrics", "firstLine", "lyricist", "composer", "translator", "tune", "key", "meter", "source"):
                if not keep.get(f) and s.get(f):
                    keep[f] = s[f]
            if s.get("lyrics") and keep.get("lyrics") and s["lyrics"].strip() != keep["lyrics"].strip():
                keep["lyrics"] = keep["lyrics"].rstrip() + "\n\n【合并自 " + s["title"] + "】\n" + s["lyrics"].strip()
            keep["attachments"] = list(keep.get("attachments") or []) + list(s.get("attachments") or [])
            keep["themes"] = sorted(set(keep.get("themes") or []) | set(s.get("themes") or []))
            keep["scenarios"] = sorted(set(keep.get("scenarios") or []) | set(s.get("scenarios") or []))
            keep["musicTypes"] = sorted(set(keep.get("musicTypes") or []) | set(s.get("musicTypes") or []))
            keep["updatedAt"] = now_str()
            keep["comment"] = (keep.get("comment") or "") + f"；已合并 {s['title']}"
            s["status"] = "merged"
            s["dupGroup"] = None
            s["dupMatches"] = []
            s["flags"] = ["已合并"]
            s["updatedAt"] = now_str()
        keep["dupGroup"] = None
        keep["dupMatches"] = []
        keep["dupResolved"] = True
        keep["updatedAt"] = now_str()
        classify_song(keep)
        song_flags(keep)
    else:
        for s in members:
            s["dupGroup"] = None
            s["dupMatches"] = []
            s["dupResolved"] = True
            s["updatedAt"] = now_str()
            song_flags(s)
    store.save()
    return {"ok": True}


# ---------------------------------------------------------------- 导入
def handle_import(store, files, uploader=""):
    """files: [(filename, bytes)] → (summary)"""
    batch = f"B{store.next_seq():04d}"
    created = []
    per_file = []
    warnings_all = []
    for filename, data in files:
        try:
            songs, warns, attach = parsers.parse_upload(filename, data, str(UPLOADS),
                                               store.data.get("settings", {}))
        except Exception as e:
            traceback.print_exc()
            per_file.append({"file": filename, "ok": 0, "needsReview": 0, "duplicates": 0, "incomplete": 0, "errors": [str(e)]})
            continue
        per_file.append({"file": filename, "songs": len(songs), "warnings": warns})
        warnings_all.extend(warns)
        if not songs:
            # 图片/音频等无法自动识别 → 建立一条待人工确认记录
            if attach["kind"] in ("image", "audio"):
                title = re.sub(r"\.[^.]+$", "", attach["name"])
                note = "图片待OCR/人工补充" if attach["kind"] == "image" else "音频暂不支持离线转写，请人工补充"
                s = default_song(title=title, firstLine="", lyrics="", source=filename,
                                 importBatch=batch, importSource=filename, ocrNote=note,
                                 attachments=[attach], status="pending")
                s["id"] = f"S{store.next_seq():05d}"
                s["number"] = ""
                s["needsReview"] = True
                created.append(s)
            continue
        for raw in songs:
            num = raw.pop("number", "")
            src = raw.pop("source", "") or filename
            s = default_song(**raw, source=src, importBatch=batch,
                            importSource=filename, attachments=[attach] if attach["kind"] in ("image", "audio", "pdf") else [])
            s["id"] = f"S{store.next_seq():05d}"
            s["number"] = num or ""
            created.append(s)

    for s in created:
        if not s.get("uploader"):
            s["uploader"] = (uploader or "").strip()
        explicit_cat = s.get("category") or ""
        explicit_sub = s.get("subcategory") or ""
        classify_song(s)
        if explicit_cat:
            s["category"] = explicit_cat
        if explicit_sub:
            s["subcategory"] = explicit_sub
        song_flags(s)
    store.songs().extend(created)
    groups = apply_dedup(store)
    # 重新标记 flags（含重复）
    for s in created:
        song_flags(s)
    summary = {
        "batch": batch,
        "files": len(files),
        "total": len(created),
        "ok": 0, "needsReview": 0, "duplicates": 0, "incomplete": 0,
        "perFile": per_file, "warnings": warnings_all,
    }
    for s in created:
        if s.get("needsReview"):
            summary["needsReview"] += 1
        if s.get("dupGroup"):
            summary["duplicates"] += 1
        if not s.get("title") or not s.get("lyrics"):
            summary["incomplete"] += 1
        if not s.get("needsReview") and not s.get("dupGroup") and s.get("title") and s.get("lyrics"):
            summary["ok"] += 1
    imp = {"batch": batch, "time": now_str(), "files": [f[0] for f in files],
           "total": len(created), "ok": summary["ok"], "needsReview": summary["needsReview"],
           "duplicates": summary["duplicates"], "incomplete": summary["incomplete"]}
    store.data["imports"].insert(0, imp)
    store.save()
    return summary


# ---------------------------------------------------------------- 整理
def run_organize(store, mode="all", reclassify=True, dedup=True, use_ai=False):
    settings = store.data.get("settings", {})
    target = store.songs() if mode == "all" else [s for s in store.songs() if s.get("status") == "pending"]
    processed = 0
    for s in target:
        if s.get("status") == "merged":
            continue
        if reclassify:
            if use_ai and settings.get("openaiKey"):
                ai = ai_classify(s, settings)
                if ai:
                    s.update({k: v for k, v in ai.items() if k in (
                        "themes", "scenarios", "musicTypes", "difficulty", "singability")})
                    s["difficultyStars"] = "★" * int(s["difficulty"]) + "☆" * (5 - int(s["difficulty"]))
                    s["singabilityStars"] = "★" * int(s["singability"]) + "☆" * (5 - int(s["singability"]))
                    s["aiConfidence"] = 0.95
                    s["needsReview"] = False
                    s["updatedAt"] = now_str()
                else:
                    classify_song(s)
            else:
                classify_song(s)
            processed += 1
        song_flags(s)
    groups = apply_dedup(store) if dedup else []
    for s in store.songs():
        song_flags(s)
    store.save()
    return {"processed": processed, "duplicateGroups": len(groups),
            "needsReview": sum(1 for s in store.songs() if s.get("needsReview") and s.get("status") != "merged"),
            "duplicates": sum(1 for s in store.songs() if s.get("dupGroup"))}


def ai_classify(song, settings):
    """调用 OpenAI 兼容接口（可选）。失败返回 None。"""
    import urllib.request
    key = (settings.get("openaiKey") or "").strip()
    base = (settings.get("openaiBase") or "https://api.openai.com/v1").rstrip("/")
    model = settings.get("openaiModel") or "gpt-4o-mini"
    if not key:
        return None
    prompt = (
        "你是赞美诗资料整理助手。根据给定曲目信息，输出 JSON（不要输出任何其他文字）。\n"
        "可选圣经主题：" + "、".join(engine.THEMES.keys()) + "\n"
        "可选崇拜场景：" + "、".join(engine.SCENARIOS.keys()) + "\n"
        "可选音乐类型：" + "、".join(engine.MUSIC_TYPES.keys()) + "\n"
        "输出格式：{\"themes\":[...],\"scenarios\":[...],\"musicTypes\":[...],\"difficulty\":1-5,\"singability\":1-5}\n"
        "曲目信息：" + json.dumps({k: song.get(k) for k in ["title", "firstLine", "lyricist", "composer", "tune", "lyrics"]},
                                  ensure_ascii=False)
    )
    body = json.dumps({"model": model, "messages": [{"role": "user", "content": prompt}],
                       "temperature": 0.2, "response_format": {"type": "json_object"}}).encode("utf-8")
    req = urllib.request.Request(base + "/chat/completions", data=body, headers={
        "Authorization": "Bearer " + key, "Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        content = data["choices"][0]["message"]["content"]
        obj = json.loads(content)
        return {
            "themes": [t for t in obj.get("themes", []) if t in engine.THEMES][:3],
            "scenarios": [t for t in obj.get("scenarios", []) if t in engine.SCENARIOS][:3],
            "musicTypes": [t for t in obj.get("musicTypes", []) if t in engine.MUSIC_TYPES][:2],
            "difficulty": max(1, min(5, int(obj.get("difficulty", 3)))),
            "singability": max(1, min(5, int(obj.get("singability", 3)))),
        }
    except Exception:
        return None


# ---------------------------------------------------------------- 统计
def compute_stats(store):
    songs = [s for s in store.songs() if s.get("status") != "merged"]
    status_counts = {k: 0 for k in STATUSES}
    for s in songs:
        status_counts[s.get("status", "pending")] = status_counts.get(s.get("status", "pending"), 0) + 1
    theme_counts, scen_counts, type_counts = {}, {}, {}
    cat_counts, sub_counts, up_counts = {}, {}, {}
    for s in songs:
        for t in s.get("themes") or []:
            theme_counts[t] = theme_counts.get(t, 0) + 1
        for t in s.get("scenarios") or []:
            scen_counts[t] = scen_counts.get(t, 0) + 1
        for t in s.get("musicTypes") or []:
            type_counts[t] = type_counts.get(t, 0) + 1
        if s.get("category"):
            cat_counts[s["category"]] = cat_counts.get(s["category"], 0) + 1
        if s.get("subcategory"):
            sub_counts[s["subcategory"]] = sub_counts.get(s["subcategory"], 0) + 1
        if s.get("uploader"):
            up_counts[s["uploader"]] = up_counts.get(s["uploader"], 0) + 1
    dup_groups = len({s["dupGroup"] for s in songs if s.get("dupGroup")})
    rated = [s.get("rating") for s in songs if s.get("rating")]
    return {
        "total": len(songs),
        "merged": sum(1 for s in store.songs() if s.get("status") == "merged"),
        "status": status_counts,
        "needsReview": sum(1 for s in songs if s.get("needsReview")),
        "duplicates": sum(1 for s in songs if s.get("dupGroup")),
        "duplicateGroups": dup_groups,
        "incomplete": sum(1 for s in songs if not s.get("title") or not s.get("lyrics")),
        "avgRating": round(sum(rated) / len(rated), 2) if rated else 0,
        "themes": sorted(theme_counts.items(), key=lambda x: -x[1])[:12],
        "scenarios": sorted(scen_counts.items(), key=lambda x: -x[1])[:10],
        "types": sorted(type_counts.items(), key=lambda x: -x[1])[:10],
        "imports": store.data.get("imports", [])[:8],
        "categories": sorted(cat_counts.items(), key=lambda x: -x[1]),
        "subcategories": sorted(sub_counts.items(), key=lambda x: -x[1])[:10],
        "uploaders": sorted(up_counts.items(), key=lambda x: -x[1])[:20],
    }


# ---------------------------------------------------------------- 筛选
def apply_filters(songs, q="", theme="", scenario="", mtype="", status="", rating=0,
                  source="", needs_review=False, dup=False, statuses=None, category="", subcategory="", uploader=""):
    res = [s for s in songs if s.get("status") != "merged"]
    if q:
        ql = q.strip().lower()
        res = [s for s in res if ql in (s.get("title") or "").lower()
               or ql in (s.get("firstLine") or "").lower()
               or ql in (s.get("lyricist") or "").lower()
               or ql in (s.get("composer") or "").lower()
               or ql in (s.get("tune") or "").lower()]
    if uploader:
        res = [s for s in res if s.get("uploader") == uploader]
    if category:
        res = [s for s in res if s.get("category") == category]
    if subcategory:
        res = [s for s in res if s.get("subcategory") == subcategory]
    if theme:
        res = [s for s in res if theme in (s.get("themes") or [])]
    if scenario:
        res = [s for s in res if scenario in (s.get("scenarios") or [])]
    if mtype:
        res = [s for s in res if mtype in (s.get("musicTypes") or [])]
    if status:
        res = [s for s in res if s.get("status") == status]
    if statuses:
        res = [s for s in res if s.get("status") in statuses]
    if rating:
        res = [s for s in res if (s.get("rating") or 0) >= float(rating)]
    if source:
        res = [s for s in res if source in (s.get("source") or "")]
    if needs_review:
        res = [s for s in res if s.get("needsReview")]
    if dup:
        res = [s for s in res if s.get("dupGroup")]
    return res


def sort_songs(songs, sort="number"):
    if sort == "title":
        return sorted(songs, key=lambda s: (s.get("title") or ""))
    if sort == "rating":
        return sorted(songs, key=lambda s: s.get("rating") or 0, reverse=True)
    if sort == "updated":
        return sorted(songs, key=lambda s: s.get("updatedAt") or "", reverse=True)

    def num_key(s):
        n = re.sub(r"\D", "", s.get("number") or "")
        return (0, int(n)) if n else (1, 0)
    return sorted(songs, key=num_key)


def build_import_template():
    """生成导入模板（按采集端字段：上传人/歌名/首句/歌词/作者/作曲/曲调/来源/备注/大类/细类）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    headers = ["上传人", "歌名", "首句", "歌词", "作者", "作曲", "曲调", "来源", "备注", "大类", "细类"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4B6EAF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    rows = [
        ["张三", "奇异恩典", "奇异恩典，何等甘甜", "奇异恩典，何等甘甜，我罪已得赦免；\n前我失丧，今被寻回，瞎眼今得看见。", "John Newton", "NEW BRITAIN", "NEW BRITAIN", "老赞美诗集", "示例：上传人必填，其余可留空", "", ""],
        ["李四", "平安夜", "平安夜，圣善夜", "平安夜，圣善夜，万暗中，光华射；\n照着圣母也照着圣婴，静享天赐安眠。", "Joseph Mohr", "Franz Gruber", "STILLE NACHT", "圣诞诗辑", "示例：大类/细类留空会自动分类", "救主耶稣", "降生"],
    ]
    for r in rows:
        ws.append(r)
    widths = [10, 16, 22, 40, 14, 14, 14, 14, 24, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- 导出
def build_export(store, req):
    scope = req.get("scope", "all")
    fmt = req.get("format", "xlsx")
    name = (req.get("name") or "").strip()
    songs = [s for s in store.songs() if s.get("status") != "merged"]

    if scope == "selected":
        ids = req.get("ids") or []
        songs = [s for s in songs if s["id"] in ids]
    elif scope == "status":
        songs = apply_filters(songs, status=req.get("status", ""))
    elif scope == "theme":
        songs = apply_filters(songs, theme=req.get("theme", ""))
    elif scope == "scenario":
        songs = apply_filters(songs, scenario=req.get("scenario", ""))
    elif scope == "type":
        songs = apply_filters(songs, mtype=req.get("type", ""))
    elif scope == "filter":
        f = req.get("filters") or {}
        songs = apply_filters(songs, q=f.get("q", ""), theme=f.get("theme", ""), scenario=f.get("scenario", ""),
                              mtype=f.get("type", ""), status=f.get("status", ""), rating=f.get("rating", 0),
                              source=f.get("source", ""), needs_review=f.get("needsReview", False),
                              dup=f.get("dup", False))
    elif scope == "report":
        songs = apply_filters(songs, statuses=["candidate", "shortlist", "final", "published"])
    elif scope == "all":
        pass

    if scope == "report":
        songs = sort_songs(songs, "number")
        title = name or "赞美诗编选总表"
        stats = compute_stats(store)
        subtitle = (f"共 {len(songs)} 首 ｜ 全部 {stats['total']} 首 ｜ 待审核 {stats['needsReview']} ｜ "
                    f"疑似重复 {stats['duplicateGroups']} 组 ｜ 生成时间 {now_str()}")
        excel, word, pdf, csvb = exporters.build_report(songs, title=title, subtitle=subtitle)
        out = []
        if fmt in ("xlsx", "all", ""):
            out.append(("xlsx", f"{title}.xlsx", excel))
        if fmt in ("docx", "all", ""):
            out.append(("docx", f"{title}.docx", word))
        if fmt in ("pdf", "all", ""):
            out.append(("pdf", f"{title}.pdf", pdf))
        if fmt in ("csv", "all", ""):
            out.append(("csv", f"{title}.csv", csvb))
    else:
        songs = sort_songs(songs, req.get("sort", "number"))
        title = name or f"赞美诗曲目({len(songs)}首)"
        fields = exporters.FIELDS
        if fmt == "xlsx":
            data = exporters.to_excel(songs, fields)
        elif fmt == "csv":
            data = exporters.to_csv(songs, fields)
        elif fmt == "docx":
            data = exporters.to_docx(songs, title=title, subtitle=f"共 {len(songs)} 首 · {now_str()}")
        elif fmt == "pdf":
            data = exporters.to_pdf(songs, title=title, subtitle=f"共 {len(songs)} 首 · {now_str()}")
        else:
            return {"ok": False, "msg": "不支持的格式"}
        ext = fmt
        out = [(fmt, f"{title}.{ext}", data)]

    results = []
    for ext, fname, data in out:
        fname = re.sub(r'[\\/:*?"<>|]', "_", fname)
        final = fname if fname.endswith("." + ext) else fname + "." + ext
        path = EXPORTS / final
        n = 1
        while path.exists():
            final = re.sub(r"\.([^.]+)$", f"-{n}.\\1", fname)
            path = EXPORTS / final
            n += 1
        path.write_bytes(data)
        rec = {"name": final, "path": str(path), "format": ext, "size": len(data),
               "time": now_str(), "count": len(songs)}
        store.data["exports"].insert(0, rec)
        results.append(rec)
    store.save()
    return {"ok": True, "files": results, "count": len(songs)}


# ---------------------------------------------------------------- 附件上传
def save_attachment(store, song, filename, data):
    attach_path = UPLOADS / f"{uuid.uuid4().hex[:8]}_{re.sub(r'[^\\w.\\-\\u4e00-\\u9fff]', '_', filename)}"
    attach_path.write_bytes(data)
    attach = {"name": filename, "path": str(attach_path), "ext": filename.rsplit(".", 1)[-1].lower() if "." in filename else ""}
    song.setdefault("attachments", []).append(attach)
    song["updatedAt"] = now_str()
    store.save()
    return attach


# ---------------------------------------------------------------- multipart
def parse_multipart(body, content_type):
    m = re.search(r"boundary=(?:\"([^\"]+)\"|([^;]+))", content_type or "")
    if not m:
        return {}
    boundary = (m.group(1) or m.group(2)).encode()
    delim = b"--" + boundary
    fields = {}
    files = []
    for chunk in body.split(delim):
        chunk = chunk.strip(b"\r\n")
        if not chunk or chunk == b"--":
            continue
        header, _, data = chunk.partition(b"\r\n\r\n")
        htext = header.decode("utf-8", "replace")
        name_m = re.search(r'name="([^"]+)"', htext)
        fname_m = re.search(r'filename="([^"]*)"', htext)
        if not name_m:
            continue
        fname = fname_m.group(1) if fname_m else ""
        if fname:
            files.append((fname, data))
        else:
            fields[name_m.group(1)] = data.decode("utf-8", "replace")
    return {"fields": fields, "files": files}


# ---------------------------------------------------------------- HTTP
class Handler(BaseHTTPRequestHandler):
    server_version = "HymnCenter/1.0"

    def log_message(self, fmt, *args):
        sys.stderr.write("[hymn] %s\n" % (fmt % args))

    def _send(self, code, body=b"", ctype="application/octet-stream", headers=None):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        for k, v in (headers or {}).items():
            self.send_header(k, v)
        self.end_headers()
        if body:
            self.wfile.write(body)

    def _json(self, obj, code=200):
        self._send(code, json.dumps(obj, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8")

    def _read_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length > 200 * 1024 * 1024:
            raise ValueError("文件过大")
        return self.rfile.read(length) if length else b""

    def _json_body(self):
        raw = self._read_body()
        return json.loads(raw.decode("utf-8")) if raw else {}

    # ---- 路由 ----
    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        qs = urllib.parse.parse_qs(parsed.query)
        try:
            if _admin_enabled() and not _is_public(path, "GET") and not _check_auth(self.headers, qs):
                return self._json({"ok": False, "msg": "需要登录"}, 401)
            if path == "/" or path == "/index.html":
                return self._serve_file(STATIC / "index.html", "text/html; charset=utf-8")
            if path in ("/mobile", "/m", "/mobile.html"):
                return self._serve_file(STATIC / "mobile.html", "text/html; charset=utf-8")
            if path in ("/qr", "/phone"):
                return self._serve_file(STATIC / "qr.html", "text/html; charset=utf-8")
            if path.startswith("/static/"):
                return self._serve_file(STATIC / path[len("/static/"):].lstrip("/"))
            if path.startswith("/files/samples/"):
                return self._serve_file(SAMPLES / Path(urllib.parse.unquote(path[len("/files/samples/"):])).name)
            if path.startswith("/files/uploads/"):
                return self._serve_file(UPLOADS / Path(urllib.parse.unquote(path[len("/files/uploads/"):])).name)
            if path.startswith("/files/exports/"):
                return self._serve_file(EXPORTS / Path(urllib.parse.unquote(path[len("/files/exports/"):])).name)
            if path == "/api/bootstrap":
                st = compute_stats(store)
                return self._json({"stats": st, "categories": {
                    "themes": list(engine.THEMES.keys()), "scenarios": list(engine.SCENARIOS.keys()),
                    "types": list(engine.MUSIC_TYPES.keys()), "statuses": STATUS_LABELS,
                    "hymnbook": CATEGORIES},
                    "settings": {**store.data.get("settings", {}), "openaiKey": mask_key(store.data.get("settings", {}).get("openaiKey", ""))},
                    "samples": [f"/files/samples/{p}" for p in sorted(os.listdir(SAMPLES)) if not p.startswith(".")],
                    "ocrAvailable": parsers.ocr_available(),
                    "ocrEngine": parsers.ocr_engine_name(),
                    "ocrEngines": [n for n in ("Vision", "tesseract.js", "tesseract")
                                   if (n == "Vision" and parsers.ensure_ocr_tool())
                                   or (n == "tesseract.js" and parsers._tessjs_ready())
                                   or (n == "tesseract" and shutil.which("tesseract"))],
                    "lanUrls": [f"http://{ip}:{PORT}" for ip in get_lan_ips()],
                    "host": HOST, "port": PORT})
            if path == "/api/stats":
                return self._json({"stats": compute_stats(store)})
            if path == "/api/songs":
                return self._list_songs(qs)
            if path == "/api/songs/template":
                return self._send(200, build_import_template(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if path == "/api/imports":
                return self._json({"imports": store.data.get("imports", [])})
            if path == "/api/duplicates":
                groups = engine.find_duplicates([s for s in store.songs() if not s.get("dupResolved")])
                return self._json({"groups": groups})
            if path == "/api/exports":
                return self._json({"exports": store.data.get("exports", [])})
            if path == "/api/auth/check":
                return self._json({"ok": True, "authed": _check_auth(self.headers, qs)})
            m = re.fullmatch(r"/api/songs/([^/]+)", path)
            if m:
                s = store.find(m.group(1))
                if not s:
                    return self._json({"ok": False, "msg": "未找到曲目"}, 404)
                return self._json({"song": s})
            self._json({"ok": False, "msg": "404"}, 404)
        except Exception as e:
            traceback.print_exc()
            self._json({"ok": False, "msg": str(e)}, 500)

    def _list_songs(self, qs):
        def gv(k, d=""):
            return qs.get(k, [d])[0]
        songs = apply_filters(
            store.songs(),
            q=gv("q"), theme=gv("theme"), scenario=gv("scenario"), mtype=gv("type"),
            status=gv("status"), rating=gv("rating", 0), source=gv("source"),
            needs_review=gv("needsReview", "0") == "1", dup=gv("dup", "0") == "1",
            statuses=(gv("statuses", "").split(",") if gv("statuses", "") else None),
            category=gv("category"), subcategory=gv("subcategory"),
            uploader=gv("uploader"),
        )
        songs = sort_songs(songs, gv("sort", "number"))
        def _int(v, d):
            try:
                return int(float(v))
            except (TypeError, ValueError):
                return d
        page = max(1, _int(gv("page", "1"), 1))
        size = min(500, max(1, _int(gv("size", "50"), 50)))
        total = len(songs)
        items = songs[(page - 1) * size: page * size]
        return self._json({"items": items, "total": total, "page": page, "size": size})

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        try:
            if _admin_enabled() and not _is_public(path, "POST") and not _check_auth(self.headers):
                return self._json({"ok": False, "msg": "需要登录"}, 401)
            if path == "/api/import/rows":
                body = self._json_body()
                rows = body.get("rows") or []
                uploader_default = (body.get("uploader") or "").strip()
                created = []
                batch = f"B{store.next_seq():04d}"
                for raw in rows:
                    if not (raw.get("title") or raw.get("lyrics") or raw.get("firstLine")):
                        continue
                    s = default_song(**{k: raw.get(k) for k in (
                        "title", "firstLine", "lyricist", "composer", "translator", "tune",
                        "key", "meter", "source", "comment", "lyrics", "number")})
                    s["id"] = f"S{store.next_seq():05d}"
                    s["uploader"] = ((raw.get("uploader") or "").strip() or uploader_default)
                    explicit_cat = (raw.get("category") or "").strip()
                    explicit_sub = (raw.get("subcategory") or "").strip()
                    classify_song(s)
                    if explicit_cat:
                        s["category"] = explicit_cat
                    if explicit_sub:
                        s["subcategory"] = explicit_sub
                    s["importBatch"] = batch
                    s["importSource"] = "表格批量导入"
                    song_flags(s)
                    created.append(s)
                store.songs().extend(created)
                apply_dedup(store)
                for s in created:
                    song_flags(s)
                store.save()
                summary = {"batch": batch, "total": len(created), "ok": 0, "needsReview": 0,
                           "duplicates": 0, "incomplete": 0}
                for s in created:
                    if s.get("needsReview"):
                        summary["needsReview"] += 1
                    if s.get("dupGroup"):
                        summary["duplicates"] += 1
                    if not s.get("title") or not s.get("lyrics"):
                        summary["incomplete"] += 1
                    if not s.get("needsReview") and not s.get("dupGroup") and s.get("title") and s.get("lyrics"):
                        summary["ok"] += 1
                imp = {"batch": batch, "time": now_str(), "files": ["表格批量导入"], "total": len(created),
                       "ok": summary["ok"], "needsReview": summary["needsReview"],
                       "duplicates": summary["duplicates"], "incomplete": summary["incomplete"]}
                store.data["imports"].insert(0, imp)
                store.save()
                return self._json({"ok": True, "summary": summary})
            if path == "/api/login":
                body = self._json_body()
                if _check_password(body.get("password", "")):
                    token = _admin_token()
                    _ADMIN_TOKENS[token] = time.time() + 30 * 86400
                    self.send_response(200)
                    self.send_header("Set-Cookie", f"hymn_admin={token}; Path=/; HttpOnly; Max-Age=604800")
                    self.send_header("Content-Type", "application/json; charset=utf-8")
                    self.end_headers()
                    self.wfile.write(json.dumps({"ok": True, "token": token}, ensure_ascii=False).encode("utf-8"))
                    return
                return self._json({"ok": False, "msg": "密码错误"}, 401)
            if path == "/api/password":
                body = self._json_body()
                if not _check_password(body.get("old", "")):
                    return self._json({"ok": False, "msg": "原密码错误"}, 401)
                newp = (body.get("new") or "").strip()
                if len(newp) < 4:
                    return self._json({"ok": False, "msg": "新密码至少 4 位"}, 400)
                import hashlib
                salt = secrets.token_hex(8)
                ADMIN["salt"] = salt
                ADMIN["hash"] = hashlib.sha256((salt + newp).encode("utf-8")).hexdigest()
                _admin_path.write_text(json.dumps(ADMIN, ensure_ascii=False, indent=1), "utf-8")
                return self._json({"ok": True, "msg": "密码已修改"})
            if path == "/api/ocr":
                ctype = self.headers.get("Content-Type", "")
                parts = parse_multipart(self._read_body(), ctype)
                if not parts["files"]:
                    return self._json({"ok": False, "msg": "未收到图片"}, 400)
                fname, data = parts["files"][0]
                safe = re.sub(r"[^\w.\-\u4e00-\u9fff]+", "_", fname)
                apath = UPLOADS / f"ocr_{uuid.uuid4().hex[:8]}_{safe}"
                apath.write_bytes(data)
                attach = {"name": fname, "path": str(apath), "ext": fname.rsplit(".", 1)[-1].lower() if "." in fname else ""}
                result = {"ok": True, "attachment": attach, "text": "", "lines": [],
                          "engine": parsers.ocr_engine_name(),
                          "parsed": {"title": "", "firstLine": "", "lyrics": "", "number": "", "note": ""}}
                text, engine, ai, lines = parsers.recognize_image_full2(str(apath), store.data.get("settings", {}))
                result["engine"] = engine
                if text and text.strip():
                    result["text"] = text
                    if ai and ai.get("lyrics"):
                        result["parsed"] = ai
                    else:
                        multi = parsers.parse_ocr_text_multi(text, fname)
                        if multi:
                            first = multi[0]
                            result["songs"] = multi
                            result["parsed"] = {"title": first["title"], "firstLine": first["firstLine"],
                                                "lyrics": first["lyrics"], "number": "",
                                                "note": f"识别到 {len(multi)} 首，已提取第一首" + ("《" + first["title"] + "》" if first["title"] else "") + "，请核对后保存"}
                        else:
                            result["parsed"] = parsers.parse_ocr_plain_text(text, fname)
                else:
                    result["parsed"]["note"] = "本机 OCR 暂不可用，请手动填写（或在设置中配置 AI 接口）"
                return self._json(result)
            if path == "/api/import":
                ctype = self.headers.get("Content-Type", "")
                body = self._read_body()
                parts = parse_multipart(body, ctype)
                files = parts.get("files") or []
                if not files:
                    return self._json({"ok": False, "msg": "未收到文件"})
                up = (parts.get("fields") or {}).get("uploader", "")
                summary = handle_import(store, files, up)
                return self._json({"ok": True, "summary": summary})
            if path == "/api/songs":
                body = self._json_body()
                s = default_song(**{k: body.get(k) for k in ("title", "firstLine", "lyricist", "composer", "translator",
                    "tune", "key", "meter", "source", "comment", "lyrics", "number", "status", "rating")})
                s["id"] = f"S{store.next_seq():05d}"
                s["themes"] = body.get("themes") or []
                s["scenarios"] = body.get("scenarios") or []
                s["musicTypes"] = body.get("musicTypes") or []
                s["uploader"] = (body.get("uploader") or "").strip()
                att = body.get("attachment")
                if att and att.get("path"):
                    s["attachments"].append({"name": att.get("name") or "拍照识别", "path": att.get("path"),
                                             "ext": att.get("ext") or ""})
                classify_song(s)
                song_flags(s)
                store.songs().append(s)
                apply_dedup(store)
                for x in store.songs():
                    song_flags(x)
                store.save()
                return self._json({"ok": True, "song": s})
            if path == "/api/organize":
                body = self._json_body()
                res = run_organize(store, mode=body.get("mode", "all"), reclassify=body.get("reclassify", True),
                                   dedup=body.get("dedup", True), use_ai=body.get("useAI", False))
                return self._json({"ok": True, **res})
            if path == "/api/export":
                body = self._json_body()
                res = build_export(store, body)
                return self._json(res)
            if path == "/api/settings":
                body = self._json_body()
                settings = store.data.setdefault("settings", {})
                if "openaiKey" in body and body["openaiKey"] != mask_key(settings.get("openaiKey", "")):
                    settings["openaiKey"] = body["openaiKey"].strip()
                for k in ("openaiBase", "openaiModel", "openaiEnabled"):
                    if k in body:
                        settings[k] = body[k]
                store.save()
                return self._json({"ok": True, "settings": {**settings, "openaiKey": mask_key(settings.get("openaiKey", ""))}})
            if path == "/api/songs/bulk":
                body = self._json_body()
                ids = body.get("ids") or []
                action = body.get("action")
                songs = [store.find(i) for i in ids]
                songs = [s for s in songs if s]
                if action == "delete":
                    idset = set(ids)
                    store.data["songs"] = [s for s in store.songs() if s["id"] not in idset]
                    store.save()
                    return self._json({"ok": True, "count": len(idset)})
                if action == "status":
                    val = body.get("value", "candidate")
                    for s in songs:
                        s["status"] = val
                        s["updatedAt"] = now_str()
                    store.save()
                    return self._json({"ok": True, "count": len(songs)})
                if action == "review":
                    val = body.get("value")
                    for s in songs:
                        s["needsReview"] = bool(val)
                        s["updatedAt"] = now_str()
                        song_flags(s)
                    store.save()
                    return self._json({"ok": True, "count": len(songs)})
                return self._json({"ok": False, "msg": "未知操作"}, 400)
            m = re.fullmatch(r"/api/duplicates/([^/]+)/resolve", path)
            if m:
                body = self._json_body()
                res = resolve_duplicate(store, m.group(1), body.get("action", "keep-both"), body.get("keepId"))
                return self._json(res)
            m = re.fullmatch(r"/api/songs/([^/]+)/attachments", path)
            if m:
                s = store.find(m.group(1))
                if not s:
                    return self._json({"ok": False, "msg": "未找到曲目"}, 404)
                ctype = self.headers.get("Content-Type", "")
                parts = parse_multipart(self._read_body(), ctype)
                if not parts["files"]:
                    return self._json({"ok": False, "msg": "未收到文件"}, 400)
                fname, data = parts["files"][0]
                attach = save_attachment(store, s, fname, data)
                return self._json({"ok": True, "attachment": attach, "song": s})
            self._json({"ok": False, "msg": "404"}, 404)
        except Exception as e:
            traceback.print_exc()
            self._json({"ok": False, "msg": str(e)}, 500)

    def do_PUT(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        m = re.fullmatch(r"/api/songs/([^/]+)", path)
        if not m:
            return self._json({"ok": False, "msg": "404"}, 404)
        try:
            s = store.find(m.group(1))
            if not s:
                return self._json({"ok": False, "msg": "未找到曲目"}, 404)
            body = self._json_body()
            for k in ("title", "firstLine", "lyricist", "composer", "translator", "tune", "key",
                      "meter", "source", "comment", "lyrics", "number", "status", "rating"):
                if k in body:
                    s[k] = body[k]
            for k in ("themes", "scenarios", "musicTypes"):
                if k in body:
                    s[k] = body[k] or []
            if body.get("uploader") is not None:
                s["uploader"] = (body.get("uploader") or "").strip()
            if body.get("category") is not None:
                s["category"] = body.get("category") or ""
            if body.get("subcategory") is not None:
                s["subcategory"] = body.get("subcategory") or ""
            if "needsReview" in body:
                s["needsReview"] = bool(body["needsReview"])
            if body.get("reclassify", True):
                classify_song(s)
            s["updatedAt"] = now_str()
            song_flags(s)
            store.save()
            return self._json({"ok": True, "song": s})
        except Exception as e:
            traceback.print_exc()
            self._json({"ok": False, "msg": str(e)}, 500)

    def do_DELETE(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        m = re.fullmatch(r"/api/songs/([^/]+)", path)
        if not m:
            return self._json({"ok": False, "msg": "404"}, 404)
        try:
            sid = m.group(1)
            before = len(store.songs())
            store.data["songs"] = [s for s in store.songs() if s["id"] != sid]
            store.save()
            return self._json({"ok": True, "deleted": before - len(store.songs())})
        except Exception as e:
            self._json({"ok": False, "msg": str(e)}, 500)

    def _serve_file(self, path, ctype=None):
        path = Path(path)
        if not path.exists() or not path.is_file():
            return self._send(404, b"not found", "text/plain")
        ctype = ctype or mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        if "charset" not in ctype and (str(path).endswith(".html") or str(path).endswith(".js") or str(path).endswith(".css")):
            ctype += "; charset=utf-8"
        self._send(200, path.read_bytes(), ctype)


# ---------------------------------------------------------------- 启动
def ensure_seeded(store):
    # seeded=True：显式“已初始化”（包括用户清空数据后），不再自动灌入示例数据
    if store.data.get("seeded"):
        return False
    if store.songs():
        store.data["seeded"] = True
        store.save()
        return False
    import seed
    for s in seed.build_seed_songs():
        s["id"] = f"S{store.next_seq():05d}"
        store.songs().append(s)
    for s in store.songs():
        classify_song(s)
    apply_dedup(store)
    for s in store.songs():
        song_flags(s)
    try:
        seed.make_sample_files(str(SAMPLES))
    except Exception as e:
        print("示例文件生成失败：", e)
    store.data["seeded"] = True
    store.save()
    return True


store = Store(DB_PATH)


def main():
    seeded = ensure_seeded(store)
    if seeded:
        print("✓ 已写入示例数据（" + str(len(store.songs())) + " 首）")
    bind_host = "0.0.0.0" if HOST == "0.0.0.0" else HOST
    server = ThreadingHTTPServer((bind_host, PORT), Handler)
    url = f"http://127.0.0.1:{PORT}"
    print()
    print("  ┌──────────────────────────────────────────────┐")
    print("  │     赞美诗资料智能整理中心 · 本地版          │")
    print(f"  │   {url}                      │")
    print("  └──────────────────────────────────────────────┘")
    if HOST == "0.0.0.0":
        print("  📱 手机访问（同一 WiFi 下打开）:")
        for ip in get_lan_ips():
            print(f"     http://{ip}:{PORT}")
        print("     若手机无法访问，请检查 Mac 防火墙是否放行 Python。")
    print("  按 Ctrl+C 停止服务")
    try:
        import webbrowser
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    except Exception:
        pass
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止。")


if __name__ == "__main__":
    main()
